"""
update_inventory.py — load RealWear inventory from NetSuite XML .xls export files.

Expects these files in the given directory (default: current dir):
  ItemQuantitySoldperMonthResults*.xls           -> monthly_sales
  CustomCurrentInventorySnapshot*.xls           -> inventory_snapshot  +  skus (supplier)
  InventoryValuationSummary*.xls                 -> inventory_valuation +  skus (desc, cost)
  PurchaseOrderHistory*.xls                      -> po_history
  Bryant_sOpenPurchaseOrders*.xls                -> open_pos
  NetSuite_Transfer_Orders_Abiertas*.xls         -> open_transfer_orders

Usage:
    python update_inventory.py "C:\\Users\\swinn\\Desktop\\Accessory Invt Mngt"
"""

import os
import sys
import glob
import unicodedata
from datetime import date, datetime
from lxml import etree
from dotenv import load_dotenv
from supabase import create_client, Client

load_dotenv()

SUPABASE_URL = os.environ["SUPABASE_URL"]
SUPABASE_SERVICE_KEY = os.environ["SUPABASE_SERVICE_KEY"]

supabase: Client = create_client(SUPABASE_URL, SUPABASE_SERVICE_KEY)
today = date.today().isoformat()

SEARCH_DIR = sys.argv[1] if len(sys.argv) > 1 else "."

NS_PREFIX = "{urn:schemas-microsoft-com:office:spreadsheet}"


# ── XML / file helpers ────────────────────────────────────────────────────────

def parse_xls_xml(path: str) -> list[list[str]]:
    """Parse a NetSuite XML SpreadsheetML .xls into a list of dense string rows."""
    with open(path, "rb") as f:
        content = f.read()
    parser = etree.XMLParser(recover=True, encoding="utf-8")
    root = etree.fromstring(content, parser=parser)

    ws = root.findall(f"{NS_PREFIX}Worksheet")[0]
    table = ws.find(f"{NS_PREFIX}Table")
    if table is None:
        return []

    result = []
    for row_el in table.findall(f"{NS_PREFIX}Row"):
        cells = row_el.findall(f"{NS_PREFIX}Cell")
        row_data: dict[int, str] = {}
        col = 0
        for c in cells:
            idx = c.get(f"{NS_PREFIX}Index")
            if idx:
                col = int(idx) - 1  # 1-based -> 0-based
            d = c.find(f"{NS_PREFIX}Data")
            row_data[col] = (d.text or "").strip() if d is not None else ""
            col += 1
        max_col = max(row_data.keys()) if row_data else -1
        result.append([row_data.get(i, "") for i in range(max_col + 1)])

    return result


def find_file(pattern: str) -> str | None:
    matches = [
        m for m in glob.glob(os.path.join(SEARCH_DIR, pattern))
        if not os.path.basename(m).startswith("~$")
    ]
    return max(matches, key=os.path.getmtime) if matches else None


def safe_int(v) -> int | None:
    try:
        return int(float(v)) if v and str(v).strip() not in ("", "nan") else None
    except (ValueError, TypeError):
        return None


def safe_float(v) -> float | None:
    try:
        return float(v) if v and str(v).strip() not in ("", "nan") else None
    except (ValueError, TypeError):
        return None


def is_valid_sku(v: str) -> bool:
    s = str(v).strip()
    if not s or s.lower() in ("nan", "none", "", "assembly/bill of materials"):
        return False
    if ":" in s:
        return False
    if s.startswith(("EarBud", "Flash Drive")):
        return False
    return any(c.isdigit() for c in s)


def upsert(table: str, rows: list, conflict_col: str = None):
    if not rows:
        print(f"  {table}: 0 rows -- skipped")
        return
    kwargs = {"on_conflict": conflict_col} if conflict_col else {}
    supabase.table(table).upsert(rows, **kwargs).execute()
    print(f"  {table}: upserted {len(rows)} rows")


# ── Snapshot column layout (confirmed by inspection of CustomCurrentInventorySnapshot-404.xls)
# Row 6 = location header row, Row 7 = sub-header row
#   Col 73: "4 - Hong Kong"  ->  Col 75 On Hand, Col 76 On Order
#   Col 87: "6 - 3pl Portland" -> Col 89 On Hand, Col 90 On Order
#   Col 206: "Total"           -> Col 208 On Hand, Col 209 On Order

COL_HK_ONHAND  = 75
COL_HK_ONORDER = 76
COL_PDX_ONHAND = 89
COL_PDX_ONORDER = 90
COL_TOT_ONHAND  = 208
COL_TOT_ONORDER = 209


# ── Readers (data only, no DB writes) ────────────────────────────────────────

def read_valuation() -> tuple[list, dict, dict]:
    """Returns (db_rows, ppu_by_sku, desc_by_sku)."""
    path = find_file("InventoryValuationSummary*.xls")
    if not path:
        print("  InventoryValuationSummary*.xls not found -- skipping")
        return [], {}, {}

    print(f"  {os.path.basename(path)}")
    rows = parse_xls_xml(path)

    db_rows, ppu_by_sku, desc_by_sku = [], {}, {}
    # Row 6 = col headers, Row 7 = category header, Rows 8+ = data
    # Cols: 0=Item, 1=Description, 2=Inv.Value, 3=%ofInv, 4=OnHand
    for r in rows[8:]:
        sku = r[0].strip() if len(r) > 0 else ""
        if not is_valid_sku(sku):
            continue
        desc = r[1].strip() if len(r) > 1 and r[1] else None
        inv_value = safe_float(r[2]) if len(r) > 2 else None
        on_hand = safe_int(r[4]) if len(r) > 4 else None
        if desc:
            desc_by_sku[sku] = desc
        if inv_value and on_hand and on_hand > 0:
            ppu_by_sku[sku] = round(inv_value / on_hand, 4)
        db_rows.append({
            "sku": sku,
            "updated_at": today,
            "on_hand": on_hand or 0,
            "inv_value": inv_value or 0.0,
        })

    return db_rows, ppu_by_sku, desc_by_sku


def read_snapshot() -> tuple[list, dict]:
    """Returns (db_rows, supplier_by_sku)."""
    # NetSuite exports the file with or without the "Custom" prefix depending on the run
    path = find_file("CustomCurrentInventorySnapshot*.xls") \
        or find_file("CurrentInventorySnapshot*.xls")
    if not path:
        print("  CurrentInventorySnapshot*.xls not found -- skipping")
        return [], {}

    print(f"  {os.path.basename(path)}")
    rows = parse_xls_xml(path)

    db_rows, supplier_by_sku = [], {}
    # Rows 0-7 = metadata + headers; data starts at row 8
    for r in rows[8:]:
        sku = r[0].strip() if len(r) > 0 else ""
        if not is_valid_sku(sku):
            continue
        supplier = r[2].strip() if len(r) > 2 and r[2] else None
        if supplier:
            supplier_by_sku[sku] = supplier

        def col(idx):
            return safe_int(r[idx]) if idx < len(r) and r[idx] else None

        db_rows.append({
            "sku": sku,
            "updated_at": today,
            "on_hand_total":    col(COL_TOT_ONHAND) or 0,
            "on_hand_portland": col(COL_PDX_ONHAND) or 0,
            "on_hand_hk":       col(COL_HK_ONHAND)  or 0,
            "on_order":         col(COL_TOT_ONORDER) or 0,
        })

    return db_rows, supplier_by_sku


def read_monthly_sales() -> list:
    """Returns db_rows for monthly_sales."""
    path = find_file("ItemQuantitySoldperMonthResults*.xls")
    if not path:
        print("  ItemQuantitySoldperMonthResults*.xls not found -- skipping")
        return []

    print(f"  {os.path.basename(path)}")
    rows = parse_xls_xml(path)
    if not rows:
        return []

    # Row 0 = headers:  col 1=Item, col 3=MaxLastSoldDate, cols 5-17=monthly qty
    # Cols 5..17 = offset 0 (current month) through 12 months ago

    # Anchor: max Last Sold Date across all rows -> determines "current month"
    anchor: date | None = None
    for r in rows[1:]:
        raw = r[3] if len(r) > 3 else ""
        if raw:
            try:
                d = datetime.fromisoformat(raw.split("T")[0]).date()
                if anchor is None or d > anchor:
                    anchor = d
            except Exception:
                pass
    if anchor is None:
        anchor = date.today()
    current_month = anchor.replace(day=1)

    def month_str(offset: int) -> str:
        m = current_month.month - offset
        y = current_month.year
        while m <= 0:
            m += 12
            y -= 1
        return date(y, m, 1).isoformat()

    db_rows = []
    for r in rows[1:]:
        sku = r[1].strip() if len(r) > 1 else ""
        if not is_valid_sku(sku):
            continue
        for offset, col_idx in enumerate(range(5, min(18, len(r)))):
            qty = safe_int(r[col_idx])
            if qty is None:
                continue
            db_rows.append({"sku": sku, "month": month_str(offset), "qty_sold": qty})

    return db_rows


def read_po_history() -> list:
    path = find_file("PurchaseOrderHistory*.xls")
    if not path:
        print("  PurchaseOrderHistory*.xls not found -- skipping")
        return []

    print(f"  {os.path.basename(path)}")
    rows = parse_xls_xml(path)

    # Detect header row
    data_start = 0
    for i, r in enumerate(rows):
        if r and any("item" in str(v).lower() or "sku" in str(v).lower() for v in r[:3]):
            data_start = i + 1
            break

    STATUS_MAP = {
        "pending receipt": "Pending", "fully billed": "Received",
        "closed": "Received", "open": "Open", "partial": "Partial",
        "cancelled": "Cancelled", "canceled": "Cancelled",
    }

    db_rows = []
    for r in rows[data_start:]:
        sku = r[0].strip() if r else ""
        if not is_valid_sku(sku):
            continue
        po_number = r[3].strip() if len(r) > 3 and r[3] else None
        if not po_number:
            continue
        raw_status = r[4].strip() if len(r) > 4 and r[4] else "Open"
        status = STATUS_MAP.get(raw_status.lower(), raw_status)
        db_rows.append({
            "sku": sku,
            "po_number": po_number,
            "vendor": r[2].strip() if len(r) > 2 else None,
            "status": status,
            "qty_ordered": safe_int(r[5]) or 0 if len(r) > 5 else 0,
            "unit_cost": safe_float(r[6]) if len(r) > 6 else None,
            "created_at": today,
        })

    return db_rows


import re as _re
_PO_RE = _re.compile(r'^PO\d+$', _re.IGNORECASE)

OPEN_PO_SKIP = {"closed", "rejected by supervisor", "nan", "none", ""}

# Statuses accepted for open_pos rows
VALID_OPEN_PO_STATUSES = {
    "Pending Receipt",
    "Partially Received",
    "Pending Bill",
    "Pending Billing/Partially Received",
    "Approved By Supervisor/Pending Receipt",
}

def is_valid_open_po_status(status: str) -> bool:
    """Accept explicit statuses or anything containing 'Pending' or 'Partially Received'."""
    if status in VALID_OPEN_PO_STATUSES:
        return True
    sl = status.lower()
    return "pending" in sl or "partially received" in sl

# Non-data rows whose col0 text we skip outright
_SKIP_COL0 = {"Purchase Orders", "Open Inventory Purchase Orders", ""}


def _build_open_po_row(sku, vendor, po_number, status, qty, cost, date=None):
    """Build a single open_pos DB row for Bryant's/xlsx fallback parsers."""
    if not is_valid_sku(sku):
        return None
    if "total" in str(sku).lower():
        return None
    po_str = str(po_number).strip() if po_number else ""
    if not _PO_RE.match(po_str):
        return None
    if not status or str(status).strip().lower() in OPEN_PO_SKIP:
        return None
    qty_val = safe_int(qty) or 0
    unit_cost = safe_float(cost) or 0.0
    return {
        "sku":         sku.strip(),
        "po_number":   po_str.upper(),
        "vendor":      vendor.strip() if vendor else None,
        "status":      str(status).strip(),
        "qty_ordered": qty_val,
        "unit_cost":   unit_cost if unit_cost else None,
        "open_amount": round(qty_val * unit_cost, 2),
        "date":        str(date).strip() if date and str(date).strip() not in ("nan", "None", "") else None,
    }


def parse_and_upsert_open_pos(supabase, folder):
    import re, xml.etree.ElementTree as ET, os
    files = [f for f in os.listdir(folder) if f.lower().startswith('openinventorypurchaseorders') and f.endswith('.xls')]
    if not files:
        print("No OpenInventoryPurchaseOrders file found"); return 0
    newest = max(files, key=lambda f: os.path.getmtime(os.path.join(folder, f)))
    path = os.path.join(folder, newest)
    print(f"Parsing: {newest}")
    with open(path, 'rb') as f:
        content = f.read()
    ns = {'ss': 'urn:schemas-microsoft-com:office:spreadsheet'}
    root = ET.fromstring(content)
    ws = root.find('.//ss:Worksheet', ns)
    table = ws.find('ss:Table', ns)
    rows = table.findall('ss:Row', ns)
    current_po = None
    current_date = None
    results = []
    for i, row in enumerate(rows):
        cells = row.findall('ss:Cell', ns)
        vals = [cell.find('ss:Data', ns).text if cell.find('ss:Data', ns) is not None else '' for cell in cells]
        if not any(v for v in vals): continue
        v0 = str(vals[0]).strip() if vals[0] else ''
        if i < 7: continue
        if '/' in v0 and len(v0) < 12 and v0.count('/') == 2:
            current_date = v0; continue
        if re.match(r'^PO\d+$', v0):
            current_po = v0; continue
        if v0.startswith('Total') or v0.startswith('Purchase Orders'): continue
        if v0 == '' and len(vals) > 6 and vals[1]:
            try: qty_open = float(vals[11]) if vals[11] else 0
            except: qty_open = 0
            if qty_open > 0:
                results.append({
                    'po_number': current_po,
                    'po_date': current_date,
                    'vendor': vals[1],
                    'status': vals[2],
                    'sku': vals[6],
                    'qty_ordered': float(vals[8]) if vals[8] else 0,
                    'qty_received': float(vals[9]) if vals[9] else 0,
                    'qty_open': qty_open,
                    'unit_price': float(vals[12]) if vals[12] else 0,
                    'amount_remaining': float(vals[13]) if vals[13] else 0,
                })
    # Coerce float quantities to int — DB columns are integer type
    for r in results:
        for col in ('qty_ordered', 'qty_received', 'qty_open'):
            if r.get(col) is not None:
                r[col] = int(r[col])
    supabase.table('open_pos').delete().neq('po_number', '').execute()
    if results:
        supabase.table('open_pos').insert(results).execute()
    print(f"Inserted {len(results)} open PO lines")
    return len(results)


def read_open_pos():
    """Fallback wrapper kept for Bryant's XLS and xlsx paths."""
    path = find_file("Bryant'sOpenPurchaseOrders*.xls") or find_file("Bryant_sOpenPurchaseOrders*.xls")
    if not path:
        return None  # caller will try xlsx fallback

    print(f"  {os.path.basename(path)}")
    rows = parse_xls_xml(path)

    # Skip metadata rows until we hit the header
    data_start = 0
    for i, r in enumerate(rows):
        if r and any(str(v).lower() in ("item", "sku") for v in r[:2]):
            data_start = i + 1
            break

    db_rows = []
    for r in rows[data_start:]:
        if not r or len(r) < 5:
            continue
        row = _build_open_po_row(
            sku=r[0], vendor=r[2] if len(r) > 2 else None,
            po_number=r[3] if len(r) > 3 else None,
            status=r[4] if len(r) > 4 else None,
            qty=r[5] if len(r) > 5 else None,
            cost=r[6] if len(r) > 6 else None,
            date=r[7] if len(r) > 7 else None,
        )
        if row:
            db_rows.append(row)
    return db_rows


def read_open_pos_xlsx():
    """Fallback: Accessory_Inv_Mgmt_*.xlsx -> 'PO History' sheet.

    Column layout (0-indexed): SKU(0), Description(1), Vendor(2),
    PO Number(3), Status(4), Qty Ordered(5), Unit Cost(6).
    Data starts at row index 3 (rows 0-2 are title/filter/header).
    """
    import openpyxl
    matches = [m for m in glob.glob(os.path.join(SEARCH_DIR, "Accessory_Inv_Mgmt_*.xlsx"))
               if not os.path.basename(m).startswith("~$")]
    if not matches:
        print("  Accessory_Inv_Mgmt_*.xlsx not found -- skipping open POs")
        return []

    path = max(matches, key=os.path.getmtime)
    print(f"  {os.path.basename(path)} [PO History sheet]")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "PO History" not in wb.sheetnames:
        wb.close()
        print("  'PO History' sheet not found -- skipping open POs")
        return []

    ws = wb["PO History"]
    db_rows = []
    for i, r in enumerate(ws.iter_rows(values_only=True)):
        if i < 3:  # title row, filter control row, header row
            continue
        if not r or r[0] is None:
            continue
        row = _build_open_po_row(
            sku=str(r[0]), vendor=str(r[2]) if r[2] else None,
            po_number=r[3], status=r[4],
            qty=r[5], cost=r[6],
        )
        if row:
            db_rows.append(row)
    wb.close()
    return db_rows


TRANSFER_ORDER_HEADER_ALIASES = {
    "to_number": "transfer_order_number",
    "fecha": "transfer_date",
    "estado": "status",
    "ubicacion_origen": "origin_location",
    "ubicacion_destino": "destination_location",
    "item_sku_/_number": "sku",
    "item_sku_number": "sku",
    "item_number": "sku",
    "item": "sku",
    "cantidad_pedida": "qty_ordered",
    "cantidad_enviada": "qty_sent",
    "cantidad_recibida": "qty_received",
    "pendiente_de_enviar": None,
    "pendiente_de_recibir": "qty_open",
    "line_id": None,
    "transfer_date": "transfer_date",
    "po_date": "transfer_date",
}


def normalize_header(cell: str) -> str:
    value = unicodedata.normalize("NFKD", str(cell or ""))
    value = "".join(ch for ch in value if not unicodedata.combining(ch))
    return value.strip().lower().replace(" ", "_")


def map_transfer_order_header(header_map: dict, cell: str, index: int):
    normalized = normalize_header(cell)
    if normalized in TRANSFER_ORDER_HEADER_ALIASES:
        target = TRANSFER_ORDER_HEADER_ALIASES[normalized]
        if target:
            header_map[target] = index
        return

    if "sku" in normalized and normalized not in ("item", "item_number"):
        header_map["sku"] = index
    elif "transfer" in normalized and "order" in normalized:
        header_map["transfer_order_number"] = index
    elif normalized in ("to", "to_number"):
        header_map["transfer_order_number"] = index
    elif "origin" in normalized or "origen" in normalized or normalized.startswith("from") or normalized.endswith("_from"):
        header_map["origin_location"] = index
    elif "destination" in normalized or "destino" in normalized or normalized.startswith("to_") or normalized.endswith("_to"):
        header_map["destination_location"] = index
    elif "vendor" in normalized:
        header_map["vendor"] = index
    elif "status" in normalized:
        header_map["status"] = index
    elif "qty_open" in normalized or ("open" in normalized and "qty" in normalized):
        header_map["qty_open"] = index
    elif "qty_ordered" in normalized or ("ordered" in normalized and "qty" in normalized):
        header_map["qty_ordered"] = index
    elif "qty_received" in normalized or ("received" in normalized and "qty" in normalized):
        header_map["qty_received"] = index
    elif "unit" in normalized and ("price" in normalized or "cost" in normalized):
        header_map["unit_price"] = index
    elif "amount" in normalized and ("remaining" in normalized or "remaining_amount" in normalized):
        header_map["amount_remaining"] = index
    elif normalized in ("date", "transfer_date", "po_date"):
        header_map["transfer_date"] = index


def is_transfer_order_header(normalized: list[str]) -> bool:
    has_sku = any("sku" in cell or cell in ("item", "item_number") for cell in normalized)
    has_to = any("transfer" in cell or cell in ("to", "to_number") for cell in normalized)
    return has_sku and has_to


def read_transfer_orders(path: str) -> list[dict]:
    rows = parse_xls_xml(path)
    if not rows:
        return []

    header_idx = None
    header_map = {}
    for i, row in enumerate(rows):
        normalized = [normalize_header(cell) for cell in row]
        if is_transfer_order_header(normalized):
            header_idx = i
            for j, cell in enumerate(row):
                map_transfer_order_header(header_map, cell, j)
            break

    if header_idx is None:
        return []

    db_rows = []
    for row in rows[header_idx + 1:]:
        sku = str(row[header_map.get("sku", -1)]).strip() if header_map.get("sku") is not None and len(row) > header_map["sku"] else ""
        if not is_valid_sku(sku):
            continue
        transfer_order_number = str(row[header_map.get("transfer_order_number", -1)]).strip() if header_map.get("transfer_order_number") is not None and len(row) > header_map["transfer_order_number"] else None
        if not transfer_order_number:
            continue
        vendor = str(row[header_map.get("vendor", -1)]).strip() if header_map.get("vendor") is not None and len(row) > header_map["vendor"] else None
        status = str(row[header_map.get("status", -1)]).strip() if header_map.get("status") is not None and len(row) > header_map["status"] else None
        origin_location = str(row[header_map.get("origin_location", -1)]).strip() if header_map.get("origin_location") is not None and len(row) > header_map["origin_location"] else None
        destination_location = str(row[header_map.get("destination_location", -1)]).strip() if header_map.get("destination_location") is not None and len(row) > header_map["destination_location"] else None
        qty_ordered = safe_int(row[header_map.get("qty_ordered", -1)]) if header_map.get("qty_ordered") is not None and len(row) > header_map["qty_ordered"] else 0
        qty_open = safe_int(row[header_map.get("qty_open", -1)]) if header_map.get("qty_open") is not None and len(row) > header_map["qty_open"] else 0
        unit_price = safe_float(row[header_map.get("unit_price", -1)]) if header_map.get("unit_price") is not None and len(row) > header_map["unit_price"] else None
        amount_remaining = safe_float(row[header_map.get("amount_remaining", -1)]) if header_map.get("amount_remaining") is not None and len(row) > header_map["amount_remaining"] else None
        transfer_date = str(row[header_map.get("transfer_date", -1)]).strip() if header_map.get("transfer_date") is not None and len(row) > header_map["transfer_date"] else None

        db_rows.append({
            "sku": sku,
            "transfer_order_number": transfer_order_number,
            "vendor": vendor,
            "status": status,
            "origin_location": origin_location,
            "destination_location": destination_location,
            "qty_ordered": qty_ordered or 0,
            "qty_open": qty_open or 0,
            "unit_price": unit_price or 0.0,
            "amount_remaining": amount_remaining or 0.0,
            "transfer_date": transfer_date or None,
        })
    return db_rows


def read_transfer_orders_file() -> list[dict]:
    path = find_file("NetSuite_Transfer_Orders_Abiertas*.xls")
    if path:
        print(f"  {os.path.basename(path)}")
        return read_transfer_orders(path)
    path = find_file("NetSuite_Transfer_Orders_Abiertas*.xlsx")
    if path:
        print(f"  {os.path.basename(path)}")
        return read_transfer_orders_xlsx()
    print("  NetSuite_Transfer_Orders_Abiertas file not found -- skipping transfer orders")
    return []


def read_transfer_orders_xlsx() -> list[dict]:
    import openpyxl
    matches = [m for m in glob.glob(os.path.join(SEARCH_DIR, "NetSuite_Transfer_Orders_Abiertas*.xlsx"))
               if not os.path.basename(m).startswith("~$")]
    if not matches:
        return []

    path = max(matches, key=os.path.getmtime)
    print(f"  {os.path.basename(path)} [Transfer Orders XLSX]")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows = list(wb.active.iter_rows(values_only=True))
    wb.close()

    header_idx = None
    header_map = {}
    for i, row in enumerate(rows):
        normalized = [normalize_header(cell) for cell in row]
        if is_transfer_order_header(normalized):
            header_idx = i
            for j, cell in enumerate(row):
                map_transfer_order_header(header_map, cell, j)
            break

    if header_idx is None:
        return []

    db_rows = []
    for row in rows[header_idx + 1:]:
        sku = str(row[header_map.get("sku", -1)]).strip() if header_map.get("sku") is not None and len(row) > header_map["sku"] else ""
        if not is_valid_sku(sku):
            continue
        transfer_order_number = str(row[header_map.get("transfer_order_number", -1)]).strip() if header_map.get("transfer_order_number") is not None and len(row) > header_map["transfer_order_number"] else None
        if not transfer_order_number:
            continue
        vendor = str(row[header_map.get("vendor", -1)]).strip() if header_map.get("vendor") is not None and len(row) > header_map["vendor"] else None
        status = str(row[header_map.get("status", -1)]).strip() if header_map.get("status") is not None and len(row) > header_map["status"] else None
        origin_location = str(row[header_map.get("origin_location", -1)]).strip() if header_map.get("origin_location") is not None and len(row) > header_map["origin_location"] else None
        destination_location = str(row[header_map.get("destination_location", -1)]).strip() if header_map.get("destination_location") is not None and len(row) > header_map["destination_location"] else None
        qty_ordered = safe_int(row[header_map.get("qty_ordered", -1)]) if header_map.get("qty_ordered") is not None and len(row) > header_map["qty_ordered"] else 0
        qty_open = safe_int(row[header_map.get("qty_open", -1)]) if header_map.get("qty_open") is not None and len(row) > header_map["qty_open"] else 0
        unit_price = safe_float(row[header_map.get("unit_price", -1)]) if header_map.get("unit_price") is not None and len(row) > header_map["unit_price"] else None
        amount_remaining = safe_float(row[header_map.get("amount_remaining", -1)]) if header_map.get("amount_remaining") is not None and len(row) > header_map["amount_remaining"] else None
        transfer_date = str(row[header_map.get("transfer_date", -1)]).strip() if header_map.get("transfer_date") is not None and len(row) > header_map["transfer_date"] else None

        db_rows.append({
            "sku": sku,
            "transfer_order_number": transfer_order_number,
            "vendor": vendor,
            "status": status,
            "origin_location": origin_location,
            "destination_location": destination_location,
            "qty_ordered": qty_ordered or 0,
            "qty_open": qty_open or 0,
            "unit_price": unit_price or 0.0,
            "amount_remaining": amount_remaining or 0.0,
            "transfer_date": transfer_date or None,
        })
    return db_rows


def parse_and_upsert_transfer_orders(supabase, folder):
    xls_files = [f for f in os.listdir(folder) if f.lower().startswith('netsuite_transfer_orders_abiertas') and f.lower().endswith('.xls')]
    xlsx_files = [f for f in os.listdir(folder) if f.lower().startswith('netsuite_transfer_orders_abiertas') and f.lower().endswith('.xlsx')]
    if not xls_files and not xlsx_files:
        print("No NetSuite_Transfer_Orders_Abiertas file found"); return 0

    if xls_files:
        newest = max(xls_files, key=lambda f: os.path.getmtime(os.path.join(folder, f)))
        path = os.path.join(folder, newest)
        print(f"Parsing: {newest}")
        results = read_transfer_orders(path)
    else:
        newest = max(xlsx_files, key=lambda f: os.path.getmtime(os.path.join(folder, f)))
        path = os.path.join(folder, newest)
        print(f"Parsing: {newest}")
        results = read_transfer_orders_xlsx()

    if not results:
        print("  no transfer orders parsed")
        return 0

    supabase.table('open_transfer_orders').delete().neq('transfer_order_number', '').execute()
    if results:
        supabase.table('open_transfer_orders').insert(results).execute()
    print(f"Inserted {len(results)} open transfer order lines")
    return len(results)


# ── Distributor stock ─────────────────────────────────────────────────────────
# Each distributor emails their own report format. One workbook sheet per
# distributor; sheet layouts vary, so each known layout gets its own explicit
# column mapping below. Unrecognized layouts are skipped with a warning
# instead of guessed at.

DISTRIBUTOR_SKIP_MATERIAL_TYPES = {
    "vendor-provided service", "subscriptions", "software (non-physical)",
}


def _sheet_base_name(sheet_name: str) -> str:
    """'07 ScanSource 1' -> 'ScanSource'; '07 Elmark Automatyka SA' -> 'Elmark Automatyka SA'."""
    name = _re.sub(r'^\d+\s*', '', sheet_name).strip()
    name = _re.sub(r'\s+\d+$', '', name).strip()
    return name or sheet_name


def _find_distributor_header(rows: list, required_any: set) -> tuple:
    for i, r in enumerate(rows[:15]):
        if not r:
            continue
        norm = [str(c).strip().lower() if c is not None else '' for c in r]
        if any(h in norm for h in required_any):
            return i, norm
    return None, None


def _parse_sap_distributor_sheet(rows: list, header_idx: int, header: list) -> tuple:
    """SAP-style export (e.g. ScanSource): Manufacturer Part Number + On Hand Qty Unrestricted."""
    sku_col = header.index('manufacturer part number')
    qty_col = header.index('on hand qty unrestricted')
    type_col = header.index('material type') if 'material type' in header else None
    date_col = header.index('date') if 'date' in header else None

    out, max_date = [], None
    for r in rows[header_idx + 1:]:
        if not r or r[sku_col] in (None, ''):
            continue
        if type_col is not None and r[type_col] and str(r[type_col]).strip().lower() in DISTRIBUTOR_SKIP_MATERIAL_TYPES:
            continue
        if date_col is not None and r[date_col]:
            d = r[date_col].date() if hasattr(r[date_col], 'date') else None
            if d and (max_date is None or d > max_date):
                max_date = d
        out.append({'sku': str(r[sku_col]).strip(), 'qty_on_hand': safe_int(r[qty_col]) or 0})
    return out, max_date


def _parse_also_distributor_sheet(rows: list, header_idx: int, header: list) -> tuple:
    """ALSO-style export: VendorPartNumber + StockQuantity, StockDate as YYYYMMDD int."""
    sku_col = header.index('vendorpartnumber')
    qty_col = header.index('stockquantity')
    date_col = header.index('stockdate') if 'stockdate' in header else None

    out, max_date = [], None
    for r in rows[header_idx + 1:]:
        if not r or r[sku_col] in (None, ''):
            continue
        if date_col is not None and r[date_col]:
            try:
                d = datetime.strptime(str(int(r[date_col])), '%Y%m%d').date()
                if max_date is None or d > max_date:
                    max_date = d
            except (ValueError, TypeError):
                pass
        out.append({'sku': str(r[sku_col]).strip(), 'qty_on_hand': safe_int(r[qty_col]) or 0})
    return out, max_date


def _parse_simple_distributor_sheet(rows: list, header_idx: int, header: list) -> tuple:
    """Simple template (e.g. Elmark): SKU ('Product_CODE') + QTY_Available."""
    sku_col = header.index('sku')
    qty_col = header.index('qty_available')

    out = []
    for r in rows[header_idx + 1:]:
        if not r or not r[sku_col]:
            continue
        raw_sku = str(r[sku_col]).strip()
        sku = raw_sku.rsplit('_', 1)[-1] if '_' in raw_sku else raw_sku
        out.append({'sku': sku, 'qty_on_hand': safe_int(r[qty_col]) or 0})
    return out, None


def read_distributor_stock(path: str) -> list[dict]:
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    by_distributor = {}  # base_name -> (parsed_rows, max_date, sheet_name)
    for sheet_name in wb.sheetnames:
        rows = list(wb[sheet_name].iter_rows(values_only=True))
        header_idx, header = _find_distributor_header(
            rows, {'manufacturer part number', 'vendorpartnumber', 'sku'})
        if header_idx is None:
            print(f"  {sheet_name}: header not recognized -- skipping")
            continue

        if 'manufacturer part number' in header and 'on hand qty unrestricted' in header:
            parsed, max_date = _parse_sap_distributor_sheet(rows, header_idx, header)
        elif 'vendorpartnumber' in header and 'stockquantity' in header:
            parsed, max_date = _parse_also_distributor_sheet(rows, header_idx, header)
        elif 'sku' in header and 'qty_available' in header:
            parsed, max_date = _parse_simple_distributor_sheet(rows, header_idx, header)
        else:
            print(f"  {sheet_name}: unrecognized column layout -- skipping")
            continue

        base = _sheet_base_name(sheet_name)
        existing = by_distributor.get(base)
        if existing is None:
            by_distributor[base] = (parsed, max_date, sheet_name)
        elif max_date is not None and (existing[1] is None or max_date > existing[1]):
            by_distributor[base] = (parsed, max_date, sheet_name)
        else:
            print(f"  {sheet_name}: superseded by newer '{existing[2]}' data for '{base}' -- skipping")

    wb.close()

    db_rows = []
    for distributor, (parsed, max_date, sheet_name) in by_distributor.items():
        updated_at = max_date.isoformat() if max_date else today
        seen_skus = set()
        for p in parsed:
            if not p['sku'] or p['sku'] in seen_skus:
                continue
            seen_skus.add(p['sku'])
            db_rows.append({
                'distributor': distributor,
                'sku': p['sku'],
                'qty_on_hand': p['qty_on_hand'],
                'updated_at': updated_at,
            })
        print(f"  {sheet_name} -> {distributor}: {len(seen_skus)} rows")

    return db_rows


def parse_and_upsert_distributor_stock(supabase, folder):
    matches = [m for m in glob.glob(os.path.join(folder, "Inventory Q*.xlsx"))
               if not os.path.basename(m).startswith("~$")]
    if not matches:
        print("No distributor inventory file found (Inventory Q*.xlsx)")
        return 0

    path = max(matches, key=os.path.getmtime)
    print(f"Parsing: {os.path.basename(path)}")
    results = read_distributor_stock(path)
    if not results:
        print("  no distributor stock rows parsed")
        return 0

    supabase.table('distributor_stock').delete().neq('distributor', '').execute()
    supabase.table('distributor_stock').insert(results).execute()
    print(f"Inserted {len(results)} distributor stock rows")
    return len(results)


# ── Demand forecast ───────────────────────────────────────────────────────────

def build_demand_forecast(sales_rows: list) -> list:
    """Pre-compute avg_3m, avg_6m, and total_12m per SKU using calendar months."""
    from collections import defaultdict

    by_sku: dict = defaultdict(dict)
    for r in sales_rows:
        by_sku[r["sku"]][r["month"]] = r["qty_sold"]

    # Anchor = most recent month in the entire dataset
    all_months = [r["month"] for r in sales_rows]
    if not all_months:
        return []
    anchor = max(all_months)  # e.g. "2026-03-01"
    ay, am = int(anchor[:4]), int(anchor[5:7])

    def calendar_sum(month_map: dict, n: int) -> int:
        """Sum qty_sold over the last n calendar months from anchor."""
        total = 0
        for i in range(n):
            mo, yr = am - i, ay
            while mo <= 0:
                mo += 12
                yr -= 1
            key = f"{yr}-{str(mo).zfill(2)}-01"
            total += month_map.get(key, 0)
        return total

    rows = []
    for sku, month_map in by_sku.items():
        all_qty = list(month_map.values())
        rows.append({
            "sku":        sku,
            "avg_3m":     round(calendar_sum(month_map, 3) / 3, 4),
            "avg_6m":     round(calendar_sum(month_map, 6) / 6, 4),
            "total_12m":  sum(all_qty),
            "updated_at": today,
        })
    return rows


# ── Cleanup ───────────────────────────────────────────────────────────────────

def cleanup_invalid_skus():
    """Delete rows with invalid SKU formats from all tables.

    Invalid formats: contains ':', starts with 'EarBud' or 'Flash Drive'.
    Child tables are deleted before the skus parent to respect FK constraints.
    """
    child_tables = ["po_history", "inventory_valuation", "monthly_sales", "inventory_snapshot", "demand_forecast"]
    all_tables = child_tables + ["skus"]
    patterns = ["%:%", "EarBud%", "Flash Drive%"]
    total = 0
    for table in all_tables:
        for pattern in patterns:
            res = supabase.table(table).delete().like("sku", pattern).execute()
            if res.data:
                total += len(res.data)
    print(f"  cleanup_invalid_skus: {total} rows removed")


# ── main ──────────────────────────────────────────────────────────────────────

def main():
    print(f"Directory: {os.path.abspath(SEARCH_DIR)}\n")

    # Step 0: remove any rows with invalid SKU formats
    print("Cleaning up invalid SKUs...")
    cleanup_invalid_skus()
    print()

    # Step 1: read all files (no DB writes yet)
    print("Reading valuation...")
    val_rows, ppu_by_sku, desc_by_sku = read_valuation()

    print("Reading snapshot...")
    snap_rows, supplier_by_sku = read_snapshot()  # matches Custom* or plain Current*

    print("Reading monthly sales...")
    sales_rows = read_monthly_sales()

    print("Reading PO history...")
    po_rows = read_po_history()

    print("Reading transfer orders...")
    transfer_rows = read_transfer_orders_file()

    print("Reading open POs...")

    # Step 2: collect all SKUs that need a parent row in skus table
    # (open_po_rows excluded here — filtered to known SKUs after skus upsert)
    all_skus = (
        {r["sku"] for r in val_rows}
        | {r["sku"] for r in snap_rows}
        | {r["sku"] for r in sales_rows}
        | {r["sku"] for r in po_rows}
        | {r["sku"] for r in transfer_rows}
    )

    # Step 3: upsert in FK-safe order
    print("\n>> skus (FK parent)")
    sku_rows = [
        {
            "sku": sku,
            "description": desc_by_sku.get(sku),
            "supplier": supplier_by_sku.get(sku),
            "supplier_email": None,
            "lead_time_days": None,
            "moq": None,
            "unit_cost": ppu_by_sku.get(sku),
            "attach_rate": None,
        }
        for sku in sorted(all_skus)
    ]
    upsert("skus", sku_rows, conflict_col="sku")

    print("\n>> inventory_valuation")
    upsert("inventory_valuation", val_rows)

    print("\n>> inventory_snapshot")
    upsert("inventory_snapshot", snap_rows)

    print("\n>> monthly_sales")
    upsert("monthly_sales", sales_rows, conflict_col="sku,month")

    print("\n>> demand_forecast")
    forecast_rows = build_demand_forecast(sales_rows)
    upsert("demand_forecast", forecast_rows, conflict_col="sku")

    print("\n>> po_history")
    upsert("po_history", po_rows)

    print("\n>> open_pos")
    n = parse_and_upsert_open_pos(supabase, SEARCH_DIR)
    print(f"  open_pos: {n} rows")

    print("\n>> open_transfer_orders")
    m = parse_and_upsert_transfer_orders(supabase, SEARCH_DIR)
    print(f"  open_transfer_orders: {m} rows")

    print("\n>> distributor_stock")
    d = parse_and_upsert_distributor_stock(supabase, SEARCH_DIR)
    print(f"  distributor_stock: {d} rows")

    print("\nDone.")


if __name__ == "__main__":
    main()
